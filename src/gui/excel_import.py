from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from data.heroes import HERO_ROSTER, HERO_ROLES
from data.xp_table import XP_PER_LEVEL
from exceptions import ValidationError
from models.capture_run import CaptureStatus
from models.hero import Hero
from storage.database import Database
from storage.repository import CaptureRunRepository, HeroRepository, SnapshotRepository

_MAX_LEVEL = max(XP_PER_LEVEL.keys())

_HEADERS = ["Hero Name", "Role", "Level", "Current XP", "Max Level (Yes/No)"]

_FILL_HEADER = PatternFill("solid", fgColor="0C0C14")
_FILL_EVEN   = PatternFill("solid", fgColor="111120")
_FILL_ODD    = PatternFill("solid", fgColor="0C0C14")

_FONT_HEADER = Font(name="Segoe UI", bold=True, color="F4D641", size=11)
_FONT_NAME   = Font(name="Segoe UI", bold=True, color="DCDCE8", size=11)
_FONT_ROLE   = Font(name="Segoe UI", italic=True, color="888899", size=11)
_FONT_DATA   = Font(name="Segoe UI", color="DCDCE8", size=11)

_CENTER = Alignment(horizontal="center", vertical="center")


def generate_template(path: str):
    wb = openpyxl.Workbook()

    # Instructions sheet
    inst = wb.active
    inst.title = "Instructions"
    inst["A1"] = "ProfTracker — Import Template"
    inst["A3"] = "How to use:"
    inst["A4"] = "  1. Switch to the 'Heroes' sheet."
    inst["A5"] = f"  2. Fill in Level (1–{_MAX_LEVEL}) and Current XP for each hero."
    inst["A6"] = "  3. If a hero is Max Level, set 'Max Level' to Yes — Current XP will be ignored."
    inst["A7"] = "  4. Rows with blank Hero Name are skipped."
    inst["A8"] = "  5. Do not rename or reorder the columns."
    inst["A3"].font = Font(name="Segoe UI", bold=True, size=11)
    for r in range(1, 9):
        inst.row_dimensions[r].height = 18

    # Heroes sheet
    ws = wb.create_sheet("Heroes")

    # Header row
    for col, label in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 22

    # Data validation
    level_dv = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2=str(_MAX_LEVEL),
        showErrorMessage=True,
        errorTitle="Invalid Level",
        error=f"Level must be between 1 and {_MAX_LEVEL}.",
    )
    maxlv_dv = DataValidation(
        type="list", formula1='"Yes,No"',
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Enter Yes or No.",
    )
    ws.add_data_validation(level_dv)
    ws.add_data_validation(maxlv_dv)

    # Pre-populate every hero at level 1
    for i, hero_name in enumerate(HERO_ROSTER):
        row = i + 2
        fill = _FILL_EVEN if i % 2 == 0 else _FILL_ODD

        c_name = ws.cell(row=row, column=1, value=hero_name)
        c_name.fill = fill; c_name.font = _FONT_NAME

        c_role = ws.cell(row=row, column=2, value=HERO_ROLES.get(hero_name, "Unknown"))
        c_role.fill = fill; c_role.font = _FONT_ROLE

        c_lvl = ws.cell(row=row, column=3, value=1)
        c_lvl.fill = fill; c_lvl.font = _FONT_DATA; c_lvl.alignment = _CENTER
        level_dv.add(c_lvl)

        c_xp = ws.cell(row=row, column=4, value=0)
        c_xp.fill = fill; c_xp.font = _FONT_DATA; c_xp.alignment = _CENTER

        c_max = ws.cell(row=row, column=5, value="No")
        c_max.fill = fill; c_max.font = _FONT_DATA; c_max.alignment = _CENTER
        maxlv_dv.add(c_max)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A2"

    wb.save(path)


def import_from_excel(path: str, db: Database) -> tuple[int, list[str]]:
    """Returns (success_count, error_messages)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Heroes"] if "Heroes" in wb.sheetnames else wb.active

    run_repo  = CaptureRunRepository(db)
    hero_repo = HeroRepository(db)
    snap_repo = SnapshotRepository(db)

    run = run_repo.create()
    now = datetime.now(timezone.utc).isoformat()
    success, errors = 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()
        if name not in HERO_ROSTER:
            errors.append(f"Row {row_idx}: unknown hero '{name}'")
            continue

        try:
            level = int(row[2]) if row[2] is not None else 1
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx} ({name}): invalid level '{row[2]}'")
            continue

        max_raw = str(row[4]).strip().lower() if row[4] is not None else "no"
        is_max  = max_raw in ("yes", "true", "1")

        if is_max or level >= _MAX_LEVEL:
            is_max, level, xp, xp_required = True, _MAX_LEVEL, 0, 0
        else:
            try:
                xp = int(row[3]) if row[3] is not None else 0
            except (TypeError, ValueError):
                xp = 0
            xp_required = XP_PER_LEVEL.get(level, 0)

        hero = Hero(
            name=name, role=HERO_ROLES.get(name, "Unknown"),
            level=level, xp=xp, xp_required=xp_required,
            is_max_level=is_max, capture_run_id=run.id, updated_at=now,
        )
        try:
            hero.validate()
        except ValidationError as e:
            errors.append(f"Row {row_idx} ({name}): {e}")
            continue

        hero_repo.upsert(hero)
        snap_repo.insert(hero)
        success += 1

    run_repo.update_status(run.id, CaptureStatus.COMPLETED, hero_count=success)
    return success, errors
